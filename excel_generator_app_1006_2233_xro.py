# 代码生成时间: 2025-10-06 22:33:38
import os
from sanic import Sanic
from sanic.response import json, file
from openpyxl import Workbook
from openpyxl.styles import Alignment, NamedStyle
from openpyxl.utils import get_column_letter


# Initialize the Sanic app
# 改进用户体验
app = Sanic("ExcelGeneratorApp")


# Define a named style for headers
HEADER_STYLE = NamedStyle(name='header',
                    font={'bold': True},
                    alignment=Alignment(horizontal='center'))


@app.route("/generate", methods=["GET"])
async def generate_excel(request):
    # Retrieve parameters from query string
    sheet_name = request.args.get("sheet", "Sheet")
# 添加错误处理
    data = request.args.get("data", "")
    
    # Check if necessary parameters are provided
    if not data:
        return json({
            "error": "No data provided"
        }, status=400)
    
    try:
        # Create a new workbook and select the active worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
# NOTE: 重要实现细节
        
        # If data is provided, split it by new lines and process each row
        rows = data.split("
")
        header = rows[0].split(",")
# FIXME: 处理边界情况
        for i, row in enumerate(rows):
            if i == 0:
                # Apply header style
                for cell in ws[1]:
                    cell.style = HEADER_STYLE
            else:
                # Write data to the worksheet
                values = row.split(",")
                ws.append(values)
                
        # Save the workbook to a temporary file
        temp_file = "temp_{}.xlsx".format(os.urandom(24).hex())
# NOTE: 重要实现细节
        wb.save(temp_file)
        wb.close()
# 改进用户体验
        
        # Return the file to the client
        return file(temp_file)
    except Exception as e:
        # Handle any unexpected errors
        return json({"error": str(e)}), 500


if __name__ == '__main__':
    # Run the app
# NOTE: 重要实现细节
    app.run(host='0.0.0.0', port=8000, debug=True)
